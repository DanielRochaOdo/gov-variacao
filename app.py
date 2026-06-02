from __future__ import annotations

import json
import os
from io import BytesIO
from pathlib import Path
from datetime import datetime
from urllib.parse import quote

from flask import Flask, Response, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from conversores import (
    ConversionError,
    TIPO_BRADESCO_BOLETO,
    TIPO_BRADESCO_TRANSFERENCIA,
    TIPO_RETORNO,
    TIPO_VARIACAO,
    gerar_txt_por_tipo,
    listar_alertas_valor_zero_bradesco_transferencia,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024  # 8 MB
_REM_FILE_COUNTERS: dict[str, int] = {}


def _nome_arquivo_rem_bradesco_pix() -> str:
    hoje = datetime.now()
    base = hoje.strftime("%d%m%Y")
    proximo = _REM_FILE_COUNTERS.get(base, 0) + 1
    _REM_FILE_COUNTERS[base] = proximo
    sufixo = "" if proximo == 1 else str(proximo - 1)
    return f"{base}{sufixo}.rem"


def _resumo_preview_bradesco_transferencia(conteudo_txt: str) -> tuple[int, int]:
    linhas = conteudo_txt.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    total_contas = 0
    total_centavos = 0

    for linha in linhas:
        if len(linha) < 240:
            continue
        if linha[7] != "3" or linha[13] != "A":
            continue
        total_contas += 1
        valor_centavos = int("".join(ch for ch in linha[119:134] if ch.isdigit()) or "0")
        total_centavos += valor_centavos

    return total_contas, total_centavos


def _gerar_modelo_bradesco_transferencia() -> BytesIO:
    headers = [
        "nome_favorecido",
        "numero_pagamento",
        "data_pagamento",
        "valor_pagamento",
        "tipo_inscricao_favorecido",
        "numero_inscricao_favorecido",
        "codigo_finalidade_complementar",
        "banco_favorecido",
        "agencia_favorecido",
        "conta_favorecido",
        "dv_conta_favorecido",
        "tipo_conta_recebedor",
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Modelo"
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)

    validacoes = [
        ("E2:E1048576", '"1,2"', "Use 1 para CPF ou 2 para CNPJ."),
        ("G2:G1048576", '"CC,PP"', "Use CC para conta corrente ou PP para poupanca."),
        ("L2:L1048576", '"01,02,03"', "Use 01=conta corrente, 02=conta pagamento ou 03=conta poupanca."),
    ]
    for intervalo, formula, mensagem in validacoes:
        validacao = DataValidation(type="list", formula1=formula, allow_blank=True)
        validacao.error = mensagem
        validacao.errorTitle = "Valor invalido"
        validacao.prompt = mensagem
        validacao.promptTitle = "Preenchimento"
        worksheet.add_data_validation(validacao)
        validacao.add(intervalo)

    worksheet.freeze_panes = "A2"

    help_sheet = workbook.create_sheet("HELP")
    help_sheet.append(["PASSO A PASSO"])
    help_sheet.append(["1", "Preencha a aba Modelo a partir da linha 2. Mantenha os nomes das colunas exatamente como estao."])
    help_sheet.append(["2", "Cada linha da aba Modelo representa um pagamento PIX Bradesco."])
    help_sheet.append(["3", "Use somente valores, sem formulas. Evite pontos, tracos e barras em CPF/CNPJ, banco, agencia e conta."])
    help_sheet.append(["4", "Salve em .xlsx e importe o arquivo no conversor selecionando PAGAMENTO BRADESCO > PIX."])
    help_sheet.append(["5", "Se houver erro de importacao, confira a linha indicada na mensagem e compare com as regras abaixo."])
    help_sheet.append([])
    help_sheet.append(["COLUNA", "OBRIGATORIA", "COMO PREENCHER", "EXEMPLO", "OBSERVACOES"])

    coluna_help = [
        (
            "nome_favorecido",
            "Sim",
            "Nome do recebedor/favorecido. Texto com ate 30 caracteres no CNAB.",
            "MARIA DA SILVA",
            "Evite acentos e caracteres especiais se o banco rejeitar o arquivo.",
        ),
        (
            "numero_pagamento",
            "Sim",
            "Identificador do pagamento usado como Seu Numero. Pode conter ate 20 caracteres.",
            "3950312",
            "Deve ser unico o suficiente para localizar o pagamento.",
        ),
        (
            "data_pagamento",
            "Sim",
            "Data em dd/mm/aaaa ou ddmmaaaa.",
            "29/04/2026",
            "Tambem aceita 29042026.",
        ),
        (
            "valor_pagamento",
            "Sim",
            "Valor em reais com duas casas decimais. Use virgula ou ponto decimal.",
            "6,88",
            "Nao use simbolo R$.",
        ),
        (
            "tipo_inscricao_favorecido",
            "Sim",
            "Codigo do tipo de documento do favorecido: 1=CPF, 2=CNPJ.",
            "1",
            "Qualquer outro codigo gera erro.",
        ),
        (
            "numero_inscricao_favorecido",
            "Sim",
            "CPF ou CNPJ do favorecido somente com numeros.",
            "00002122624329",
            "CPF sera completado com zeros a esquerda no CNAB.",
        ),
        (
            "codigo_finalidade_complementar",
            "Nao",
            "Tipo de conta conforme manual neste fluxo: CC=conta corrente, PP=conta poupanca.",
            "CC",
            "Se vazio, o conversor usa CC.",
        ),
        (
            "banco_favorecido",
            "Sim",
            "Codigo COMPE do banco favorecido com 3 digitos.",
            "237",
            "Neste fluxo, 001 e 341 nao sao aceitos pelo conversor.",
        ),
        (
            "agencia_favorecido",
            "Sim",
            "Agencia do favorecido com 1 a 5 digitos, sem digito verificador.",
            "02793",
            "O conversor completa com zeros a esquerda quando necessario.",
        ),
        (
            "conta_favorecido",
            "Sim",
            "Conta do favorecido somente com numeros.",
            "000000075961",
            "Informe a conta sem separadores.",
        ),
        (
            "dv_conta_favorecido",
            "Sim",
            "Digito verificador da conta. Use 1 caractere.",
            "9",
            "Pode ser numero ou letra, conforme a conta.",
        ),
        (
            "tipo_conta_recebedor",
            "Nao",
            "Codigo do tipo de conta do recebedor: 01=conta corrente, 02=conta pagamento, 03=conta poupanca.",
            "01",
            "Se vazio, o conversor tenta derivar de CC/PP.",
        ),
    ]
    for linha in coluna_help:
        help_sheet.append(list(linha))

    help_sheet.append([])
    help_sheet.append(["CODIGOS E REGRAS FIXAS DO PIX BRADESCO"])
    codigos = [
        ["forma_iniciacao", "05", "PIX por dados bancarios. Esta regra e fixa no conversor."],
        ["forma_lancamento", "45", "PIX Transferencia. Se a coluna existir e vier preenchida, deve ser 45."],
        ["codigo_finalidade_ted", "00010", "Padrao usado pelo conversor quando a coluna nao existe ou esta vazia."],
        ["codigo_finalidade_complementar", "CC ou PP", "CC=conta corrente; PP=conta poupanca."],
        ["tipo_conta_recebedor", "01, 02 ou 03", "01=conta corrente; 02=conta pagamento; 03=conta poupanca."],
        ["tipo_inscricao_favorecido", "1 ou 2", "1=CPF; 2=CNPJ."],
    ]
    for linha in codigos:
        help_sheet.append(linha)

    help_sheet["A1"].font = Font(bold=True, size=14)
    for row in help_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in help_sheet[8]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in range(1, help_sheet.max_row + 1):
        if help_sheet.cell(row=row, column=1).value == "CODIGOS E REGRAS FIXAS DO PIX BRADESCO":
            help_sheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            break

    help_sheet.column_dimensions["A"].width = 30
    help_sheet.column_dimensions["B"].width = 16
    help_sheet.column_dimensions["C"].width = 68
    help_sheet.column_dimensions["D"].width = 20
    help_sheet.column_dimensions["E"].width = 54

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/")
def index() -> str:
    return render_template("index.html")


@app.post("/api/converter")
def converter() -> Response:
    tipo = (request.form.get("tipo") or "").strip().upper()
    arquivo = request.files.get("arquivo")

    tipos_validos = {
        TIPO_RETORNO,
        TIPO_VARIACAO,
        TIPO_BRADESCO_TRANSFERENCIA,
        TIPO_BRADESCO_BOLETO,
    }
    if tipo not in tipos_validos:
        return jsonify(
            {
                "erro": (
                    "Tipo invalido. Selecione RETORNO, VARIACAO, "
                    "Pagamento Bradesco > PIX ou Pagamento Bradesco > Boleto."
                )
            }
        ), 400

    if arquivo is None or arquivo.filename is None or arquivo.filename.strip() == "":
        return jsonify({"erro": "Envie um arquivo antes de converter."}), 400

    nome_arquivo = arquivo.filename.strip()
    extensao = Path(nome_arquivo).suffix.lower()
    formatos_excel = (".xlsx", ".xlsm", ".xltx", ".xltm")
    if extensao not in formatos_excel:
        return jsonify({"erro": "Formato invalido. Envie um arquivo .xlsx."}), 400

    try:
        conteudo_txt = gerar_txt_por_tipo(arquivo.stream, tipo)
    except ConversionError as exc:
        return jsonify({"erro": str(exc)}), 400
    except Exception:
        return jsonify({"erro": "Falha inesperada ao converter o arquivo."}), 500

    if tipo == TIPO_BRADESCO_TRANSFERENCIA:
        nome_saida = _nome_arquivo_rem_bradesco_pix()
    else:
        nome_base = Path(nome_arquivo).stem
        prefixos = {
            TIPO_RETORNO: "retorno_",
            TIPO_VARIACAO: "variacao_",
            TIPO_BRADESCO_BOLETO: "bradesco_boleto_",
        }
        prefixo = prefixos.get(tipo, "arquivo_")
        nome_saida = f"{prefixo}{nome_base}.txt"

    try:
        conteudo_bytes = conteudo_txt.encode("latin-1")
        charset = "iso-8859-1"
    except UnicodeEncodeError:
        # Fallback defensivo para caracteres fora do latin-1.
        conteudo_bytes = conteudo_txt.encode("latin-1", errors="replace")
        charset = "iso-8859-1"

    headers = {"Content-Disposition": f'attachment; filename="{nome_saida}"'}
    if tipo == TIPO_BRADESCO_TRANSFERENCIA:
        contas_total, valor_total_centavos = _resumo_preview_bradesco_transferencia(conteudo_txt)
        headers["X-Preview-Contas"] = str(contas_total)
        headers["X-Preview-Valor-Centavos"] = str(valor_total_centavos)

        alertas = listar_alertas_valor_zero_bradesco_transferencia(conteudo_txt)
        total_alertas = len(alertas)
        if total_alertas > 0:
            max_alertas_header = 50
            alertas_expostos = alertas[:max_alertas_header]
            payload = json.dumps(alertas_expostos, ensure_ascii=False, separators=(",", ":"))
            headers["X-Valor-Zero-Total"] = str(total_alertas)
            headers["X-Valor-Zero-Truncated"] = "1" if total_alertas > max_alertas_header else "0"
            headers["X-Valor-Zero-Details"] = quote(payload, safe="")

    return Response(
        conteudo_bytes,
        content_type=f"text/plain; charset={charset}",
        headers=headers,
    )


@app.get("/api/modelo/bradesco-transferencia")
def baixar_modelo_bradesco_transferencia() -> Response:
    return send_file(
        _gerar_modelo_bradesco_transferencia(),
        as_attachment=True,
        download_name="modelo_bradesco_transferencia.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.errorhandler(413)
def payload_too_large(_: Exception):
    return jsonify({"erro": "Arquivo muito grande. Limite: 8 MB."}), 413


if __name__ == "__main__":
    app.run(
        host=os.getenv("HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", "5000")),
        debug=os.getenv("DEBUG", "true").lower() in {"1", "true", "yes", "on"},
    )
