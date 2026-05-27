from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BufferedIOBase, BytesIO
import math
import re
from typing import Any, Dict, List, Mapping
import unicodedata

from openpyxl import load_workbook


class ConversionError(Exception):
    """Erro de validacao ou formatacao durante a conversao."""


TIPO_RETORNO = "RETORNO"
TIPO_VARIACAO = "VARIACAO"
TIPO_BRADESCO_TRANSFERENCIA = "BRADESCO_TRANSFERENCIA"
TIPO_BRADESCO_BOLETO = "BRADESCO_BOLETO"

TIPOS_SUPORTADOS = {
    TIPO_RETORNO,
    TIPO_VARIACAO,
    TIPO_BRADESCO_TRANSFERENCIA,
    TIPO_BRADESCO_BOLETO,
}

NOMES_EXIBICAO = {
    TIPO_RETORNO: "RETORNO",
    TIPO_VARIACAO: "VARIACAO",
    TIPO_BRADESCO_TRANSFERENCIA: "PAGAMENTO BRADESCO > PIX",
    TIPO_BRADESCO_BOLETO: "PAGAMENTO BRADESCO > BOLETO",
}


def gerar_txt_por_tipo(excel_file: BufferedIOBase | BytesIO, tipo: str) -> str:
    tipo_normalizado = (tipo or "").strip().upper()
    if tipo_normalizado == TIPO_RETORNO:
        return gerar_layout_retorno(excel_file)
    if tipo_normalizado == TIPO_VARIACAO:
        return gerar_layout_variacao(excel_file)
    if tipo_normalizado == TIPO_BRADESCO_TRANSFERENCIA:
        return gerar_layout_bradesco_transferencia(excel_file)
    if tipo_normalizado == TIPO_BRADESCO_BOLETO:
        return gerar_layout_bradesco_boleto(excel_file)
    raise ConversionError(
        "Tipo invalido. Use RETORNO, VARIACAO, BRADESCO_TRANSFERENCIA ou BRADESCO_BOLETO."
    )


def ajustar_txt_bradesco_por_tipo(txt_file: BufferedIOBase | BytesIO, tipo: str) -> str:
    tipo_normalizado = (tipo or "").strip().upper()
    if tipo_normalizado not in {TIPO_BRADESCO_TRANSFERENCIA, TIPO_BRADESCO_BOLETO}:
        raise ConversionError("Ajuste de TXT disponivel apenas para os tipos Bradesco.")

    if hasattr(txt_file, "seek"):
        txt_file.seek(0)
    if hasattr(txt_file, "read"):
        bruto = txt_file.read()
    elif isinstance(txt_file, (bytes, bytearray)):
        bruto = bytes(txt_file)
    else:
        raise ConversionError("Formato de entrada TXT invalido.")
    if isinstance(bruto, str):
        texto = bruto
    else:
        texto = _decodificar_texto(bruto)

    linhas_raw = texto.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if linhas_raw and linhas_raw[-1] == "":
        linhas_raw = linhas_raw[:-1]
    if not linhas_raw:
        raise ConversionError("Arquivo TXT vazio.")

    linhas = [_normalizar_linha_240(linha) for linha in linhas_raw]

    for idx, linha in enumerate(linhas):
        if _tipo_registro(linha) == "0":
            linha = _set_slice(linha, 164, 166, "089")
            linha = _corrigir_alinhamento_header_bradesco(linha)
        if _tipo_registro(linha) == "1":
            layout_lote = "045" if tipo_normalizado == TIPO_BRADESCO_TRANSFERENCIA else "040"
            linha = _set_slice(linha, 14, 16, layout_lote)
            linha = _corrigir_alinhamento_header_bradesco(linha)
        linhas[idx] = _normalizar_linha_240(linha)

    linhas = _corrigir_trailers_bradesco(linhas, tipo_normalizado)
    return "\n".join(linhas) + "\n"


def listar_alertas_valor_zero_bradesco_transferencia(conteudo_txt: str) -> List[Dict[str, str]]:
    alertas: List[Dict[str, str]] = []
    linhas = conteudo_txt.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if linhas and linhas[-1] == "":
        linhas = linhas[:-1]

    for idx, linha_raw in enumerate(linhas):
        linha = _normalizar_linha_240(linha_raw)
        if _tipo_registro(linha) != "3":
            continue
        if len(linha) < 14 or linha[13] != "A":
            continue

        valor_raw = linha[119:134]
        valor_centavos = int(_somente_digitos(valor_raw) or "0")
        if valor_centavos != 0:
            continue

        dv_ag = linha[28:29].strip()
        dv_cc = linha[41:42].strip()
        conta_fmt = linha[29:41].strip()
        if dv_cc:
            conta_fmt = f"{conta_fmt}-{dv_cc}"
        agencia_fmt = linha[23:28].strip()
        if dv_ag:
            agencia_fmt = f"{agencia_fmt}-{dv_ag}"

        alertas.append(
            {
                "linha_arquivo": str(idx + 1),
                "registro_lote": linha[8:13].strip() or "00000",
                "banco_favorecido": linha[20:23].strip() or "000",
                "agencia_favorecido": agencia_fmt,
                "conta_favorecido": conta_fmt,
                "nome_favorecido": linha[43:73].strip(),
                "valor_pagamento": "0,00",
            }
        )

    return alertas


def gerar_layout_retorno(excel_file: BufferedIOBase | BytesIO) -> str:
    linhas = [
        _formatar_linha_retorno(row)
        for row in _ler_planilha_excel(
            excel_file,
            required_headers=[
                "ano_mes",
                "orgao",
                "matricula",
                "consignataria",
                "valor_parcela",
                "cpf",
                "contrato",
                "nome_servidor",
            ],
            nome_layout="RETORNO",
        )
    ]
    return "\n".join(linhas) + ("\n" if linhas else "")


def gerar_layout_variacao(excel_file: BufferedIOBase | BytesIO) -> str:
    linhas = [
        _formatar_linha_variacao(row)
        for row in _ler_planilha_excel(
            excel_file,
            required_headers=[
                "consignataria",
                "nome_consignataria",
                "instituicao",
                "contrato",
                "nome_servidor",
                "cpf",
                "orgao",
                "matricula",
                "tipo_ajuste",
                "categoria_ajuste",
                "data_inicial",
                "valor_total",
                "qtd_parcelas",
                "valor_parcela",
                "proxima_parcela",
            ],
            nome_layout="VARIACAO",
        )
    ]
    return "\n".join(linhas) + ("\n" if linhas else "")


def _decodificar_texto(bruto: bytes) -> str:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            return bruto.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ConversionError("Nao foi possivel decodificar o TXT enviado.")


def _normalizar_linha_240(linha: str) -> str:
    if linha is None:
        linha = ""
    linha = linha.rstrip("\n").rstrip("\r")
    if len(linha) > 240:
        tipo = linha[7] if len(linha) >= 8 else ""
        excesso = len(linha) - 240
        if tipo in {"0", "1"} and excesso > 0 and len(linha) >= 102 + excesso:
            # Corrige desalinhamento comum por nome da empresa acima de 30 posicoes.
            linha = linha[:72] + linha[72:102] + linha[102 + excesso :]
        if len(linha) > 240 and all(ch == " " for ch in linha[240:]):
            linha = linha[:240]
        elif len(linha) > 240:
            linha = linha[:240]
    if len(linha) < 240:
        linha = f"{linha:<240}"
    return linha


def _tipo_registro(linha: str) -> str:
    if len(linha) < 8:
        return ""
    return linha[7]


def _set_slice(linha: str, inicio: int, fim: int, valor: str) -> str:
    largura = fim - inicio + 1
    if len(linha) < fim:
        linha = f"{linha:<{fim}}"
    campo = f"{(valor or '')[:largura]:<{largura}}"
    return linha[: inicio - 1] + campo + linha[fim:]


def _num_slice(linha: str, inicio: int, fim: int) -> int:
    digitos = _somente_digitos(linha[inicio - 1 : fim]) if len(linha) >= inicio else ""
    return int(digitos) if digitos else 0


def _valor_pagamento_linha(linha: str, tipo: str) -> int:
    if _tipo_registro(linha) != "3":
        return 0
    segmento = linha[13] if len(linha) >= 14 else ""
    if tipo == TIPO_BRADESCO_TRANSFERENCIA:
        if segmento != "A":
            return 0
        return _num_slice(linha, 120, 134)
    if tipo == TIPO_BRADESCO_BOLETO:
        if segmento != "J":
            return 0
        if len(linha) >= 19 and linha[17:19] == "52":
            return 0
        return _num_slice(linha, 153, 167)
    return 0


def _quantidade_moeda_linha(linha: str, tipo: str) -> int:
    if _tipo_registro(linha) != "3":
        return 0
    segmento = linha[13] if len(linha) >= 14 else ""
    if tipo == TIPO_BRADESCO_TRANSFERENCIA:
        if segmento != "A":
            return 0
        return _num_slice(linha, 105, 119)
    if tipo == TIPO_BRADESCO_BOLETO:
        if segmento != "J":
            return 0
        if len(linha) >= 19 and linha[17:19] == "52":
            return 0
        return _num_slice(linha, 168, 182)
    return 0


def _corrigir_trailers_bradesco(linhas: List[str], tipo: str) -> List[str]:
    linhas = [_normalizar_linha_240(linha) for linha in linhas if linha != ""]
    if not linhas:
        raise ConversionError("Arquivo TXT vazio apos normalizacao.")

    banco_base = linhas[0][0:3] if len(linhas[0]) >= 3 else "237"

    idx_trailer_arquivo = next((i for i, l in enumerate(linhas) if _tipo_registro(l) == "9"), None)
    if idx_trailer_arquivo is None:
        linhas.append(_normalizar_linha_240(f"{banco_base}99999{' ' * 232}"))
        idx_trailer_arquivo = len(linhas) - 1

    lotes = [linha[3:7] for linha in linhas if _tipo_registro(linha) == "1"]
    lotes_unicos: List[str] = []
    for lote in lotes:
        if lote not in lotes_unicos:
            lotes_unicos.append(lote)

    for lote in lotes_unicos:
        indices_lote = [i for i, linha in enumerate(linhas) if linha[3:7] == lote and _tipo_registro(linha) != "9"]
        qtd_registros_lote = len(indices_lote)
        soma_valores = sum(_valor_pagamento_linha(linhas[i], tipo) for i in indices_lote)
        soma_qtd_moeda = sum(_quantidade_moeda_linha(linhas[i], tipo) for i in indices_lote)

        idx_trailer_lote = next(
            (i for i in indices_lote if _tipo_registro(linhas[i]) == "5"),
            None,
        )
        if idx_trailer_lote is None:
            trailer_novo = _normalizar_linha_240(f"{banco_base}{lote}5{' ' * 232}")
            linhas.insert(idx_trailer_arquivo, trailer_novo)
            idx_trailer_arquivo += 1
            idx_trailer_lote = idx_trailer_arquivo - 1
            indices_lote = [i for i, linha in enumerate(linhas) if linha[3:7] == lote and _tipo_registro(linha) != "9"]
            qtd_registros_lote = len(indices_lote)

        trailer = linhas[idx_trailer_lote]
        trailer = _set_slice(trailer, 18, 23, _cnab_num(qtd_registros_lote, 6))
        trailer = _set_slice(trailer, 24, 41, _cnab_num(soma_valores, 18))
        trailer = _set_slice(trailer, 42, 59, _cnab_num(soma_qtd_moeda, 18))
        linhas[idx_trailer_lote] = _normalizar_linha_240(trailer)

    idx_trailer_arquivo = next((i for i, l in enumerate(linhas) if _tipo_registro(l) == "9"), len(linhas) - 1)
    qtd_lotes = len(lotes_unicos)
    qtd_registros_arquivo = len(linhas)
    trailer_arquivo = linhas[idx_trailer_arquivo]
    trailer_arquivo = _set_slice(trailer_arquivo, 1, 3, banco_base)
    trailer_arquivo = _set_slice(trailer_arquivo, 4, 7, "9999")
    trailer_arquivo = _set_slice(trailer_arquivo, 8, 8, "9")
    trailer_arquivo = _set_slice(trailer_arquivo, 18, 23, _cnab_num(qtd_lotes, 6))
    trailer_arquivo = _set_slice(trailer_arquivo, 24, 29, _cnab_num(qtd_registros_arquivo, 6))
    linhas[idx_trailer_arquivo] = _normalizar_linha_240(trailer_arquivo)

    return linhas


def _corrigir_alinhamento_header_bradesco(linha: str) -> str:
    linha = _normalizar_linha_240(linha)
    if _tipo_registro(linha) not in {"0", "1"}:
        return linha

    agencia_esperada = _somente_digitos(linha[52:57])
    conta_esperada = _somente_digitos(linha[58:70])
    if len(agencia_esperada) == 5 and len(conta_esperada) == 12:
        return linha

    for deslocamento in (4, 3, 2, 1):
        inicio_agencia = 52 + deslocamento
        agencia = _somente_digitos(linha[inicio_agencia : inicio_agencia + 5])
        conta = _somente_digitos(linha[inicio_agencia + 6 : inicio_agencia + 18])
        if len(agencia) == 5 and len(conta) >= 10:
            linha = linha[:52] + linha[52 + deslocamento :] + (" " * deslocamento)
            return _normalizar_linha_240(linha)

    return linha


def gerar_layout_bradesco_transferencia(excel_file: BufferedIOBase | BytesIO) -> str:
    registros = _ler_planilha_excel(
        excel_file,
        required_headers=[
            "tipo_inscricao_empresa",
            "numero_inscricao_empresa",
            "convenio",
            "agencia_empresa",
            "conta_empresa",
            "nome_empresa",
            "nome_favorecido",
            "data_pagamento",
            "valor_pagamento",
            "forma_iniciacao",
            "tipo_inscricao_favorecido",
            "numero_inscricao_favorecido",
        ],
        nome_layout=TIPO_BRADESCO_TRANSFERENCIA,
    )
    if not registros:
        raise ConversionError("Planilha vazia para BRADESCO_TRANSFERENCIA.")

    base = _base_bradesco(registros[0])
    # Transferencia Bradesco exige identificacao PIX no header do arquivo (posicoes 172-174).
    base["header_pix"] = "PIX"
    _validar_header_pix_transferencia(base)
    lote = _cnab_num(registros[0].get("lote"), 4, default="1")

    linhas: List[str] = []
    header_arquivo = _linha_header_arquivo_bradesco(base)
    _validar_posicao_pix_header_arquivo(header_arquivo)
    linhas.append(header_arquivo)
    linhas.append(_linha_header_lote_transferencia(base, lote, registros[0]))

    total_valor = 0
    total_qtd_moeda = 0
    sequencial = 1
    for row in registros:
        _validar_modalidade_pix_transferencia(row)
        valor_pagamento = _numero_escalado(row.get("valor_pagamento"), 2)
        qtd_moeda = _numero_escalado(row.get("quantidade_moeda"), 5)
        total_valor += valor_pagamento
        total_qtd_moeda += qtd_moeda

        linhas.append(_linha_segmento_a(base, lote, row, sequencial))
        sequencial += 1
        linhas.append(_linha_segmento_b(base, lote, row, sequencial))
        sequencial += 1

    qtd_registros_lote = len(registros) * 2 + 2
    linhas.append(
        _linha_trailer_lote(
            base=base,
            lote=lote,
            qtd_registros_lote=qtd_registros_lote,
            soma_valores=total_valor,
            soma_qtd_moeda=total_qtd_moeda,
            numero_aviso_debito=registros[0].get("numero_aviso_debito"),
        )
    )

    qtd_registros_arquivo = len(linhas) + 1
    linhas.append(
        _linha_trailer_arquivo(
            base=base,
            qtd_lotes=1,
            qtd_registros_arquivo=qtd_registros_arquivo,
            qtd_contas_conciliacao=registros[0].get("qtd_contas_conciliacao"),
        )
    )

    linhas = [
        _corrigir_alinhamento_header_bradesco(linha) if _tipo_registro(linha) in {"0", "1"} else linha
        for linha in linhas
    ]

    return "\r\n".join(linhas) + "\r\n"


def gerar_layout_bradesco_boleto(excel_file: BufferedIOBase | BytesIO) -> str:
    registros = _ler_planilha_excel(
        excel_file,
        required_headers=[
            "tipo_inscricao_empresa",
            "numero_inscricao_empresa",
            "convenio",
            "agencia_empresa",
            "conta_empresa",
            "nome_empresa",
            "codigo_barras",
            "nome_beneficiario",
            "data_pagamento",
            "valor_pagamento",
        ],
        nome_layout=TIPO_BRADESCO_BOLETO,
    )
    if not registros:
        raise ConversionError("Planilha vazia para BRADESCO_BOLETO.")

    base = _base_bradesco(registros[0])
    lote = _cnab_num(registros[0].get("lote"), 4, default="1")

    linhas: List[str] = []
    linhas.append(_linha_header_arquivo_bradesco(base))
    linhas.append(_linha_header_lote_boleto(base, lote, registros[0]))

    total_valor = 0
    total_qtd_moeda = 0
    sequencial = 1
    for row in registros:
        valor_pagamento = _numero_escalado(row.get("valor_pagamento"), 2)
        qtd_moeda = _numero_escalado(row.get("quantidade_moeda"), 5)
        total_valor += valor_pagamento
        total_qtd_moeda += qtd_moeda

        linhas.append(_linha_segmento_j(base, lote, row, sequencial))
        sequencial += 1
        linhas.append(_linha_segmento_j52(base, lote, row, sequencial))
        sequencial += 1

    qtd_registros_lote = len(registros) * 2 + 2
    linhas.append(
        _linha_trailer_lote(
            base=base,
            lote=lote,
            qtd_registros_lote=qtd_registros_lote,
            soma_valores=total_valor,
            soma_qtd_moeda=total_qtd_moeda,
            numero_aviso_debito=registros[0].get("numero_aviso_debito"),
        )
    )

    qtd_registros_arquivo = len(linhas) + 1
    linhas.append(
        _linha_trailer_arquivo(
            base=base,
            qtd_lotes=1,
            qtd_registros_arquivo=qtd_registros_arquivo,
            qtd_contas_conciliacao=registros[0].get("qtd_contas_conciliacao"),
        )
    )

    linhas = [
        _corrigir_alinhamento_header_bradesco(linha) if _tipo_registro(linha) in {"0", "1"} else linha
        for linha in linhas
    ]

    return "\n".join(linhas) + "\n"


def _ler_planilha_excel(
    excel_file: BufferedIOBase | BytesIO,
    required_headers: List[str] | None = None,
    nome_layout: str = "",
) -> List[Dict[str, Any]]:
    try:
        if hasattr(excel_file, "seek"):
            excel_file.seek(0)
        workbook = load_workbook(excel_file, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - depende do parser do openpyxl
        raise ConversionError("Nao foi possivel ler o arquivo Excel enviado.") from exc

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    headers_raw = next(rows, None)
    if not headers_raw:
        raise ConversionError("A planilha esta vazia.")

    headers = [_normalizar_header(valor) for valor in headers_raw]
    if not any(headers):
        raise ConversionError("Cabecalho invalido na planilha.")

    if required_headers:
        missing = [col for col in required_headers if col not in headers]
        if missing:
            layout_detectado = _detectar_layout(headers)
            dica = ""
            if layout_detectado and layout_detectado != nome_layout:
                nome_detectado = NOMES_EXIBICAO.get(layout_detectado, layout_detectado)
                nome_esperado = NOMES_EXIBICAO.get(nome_layout, nome_layout)
                dica = (
                    f" A planilha enviada parece ser do layout {nome_detectado}. "
                    f"Selecione {nome_esperado} para gerar o TXT correto."
                )
            nome_layout_exibicao = NOMES_EXIBICAO.get(nome_layout, nome_layout)
            raise ConversionError(
                f"Planilha nao corresponde ao layout {nome_layout_exibicao}. "
                f"Colunas ausentes: {', '.join(missing)}.{dica}"
            )

    itens: List[Dict[str, Any]] = []
    for linha_planilha, values in enumerate(rows, start=2):
        if values is None or all(valor is None or str(valor).strip() == "" for valor in values):
            continue
        registro = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        registro["__linha_planilha"] = linha_planilha
        itens.append(registro)

    return itens


def _normalizar_header(valor: Any) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto


def _detectar_layout(headers: List[str]) -> str:
    headers_set = set(headers)
    retorno_base = {"ano_mes", "orgao", "matricula", "consignataria", "valor_parcela", "cpf", "contrato", "nome_servidor"}
    variacao_base = {
        "consignataria",
        "nome_consignataria",
        "instituicao",
        "contrato",
        "nome_servidor",
        "cpf",
        "orgao",
        "matricula",
        "tipo_ajuste",
        "categoria_ajuste",
        "data_inicial",
        "valor_total",
        "qtd_parcelas",
        "valor_parcela",
        "proxima_parcela",
    }
    bradesco_transferencia_base = {
        "tipo_inscricao_empresa",
        "numero_inscricao_empresa",
        "convenio",
        "agencia_empresa",
        "conta_empresa",
        "nome_empresa",
        "nome_favorecido",
        "data_pagamento",
        "valor_pagamento",
        "forma_iniciacao",
        "tipo_inscricao_favorecido",
        "numero_inscricao_favorecido",
    }
    bradesco_boleto_base = {
        "tipo_inscricao_empresa",
        "numero_inscricao_empresa",
        "convenio",
        "agencia_empresa",
        "conta_empresa",
        "nome_empresa",
        "codigo_barras",
        "nome_beneficiario",
        "data_pagamento",
        "valor_pagamento",
    }
    if variacao_base.issubset(headers_set):
        return TIPO_VARIACAO
    if retorno_base.issubset(headers_set):
        return TIPO_RETORNO
    if bradesco_transferencia_base.issubset(headers_set):
        return TIPO_BRADESCO_TRANSFERENCIA
    if bradesco_boleto_base.issubset(headers_set):
        return TIPO_BRADESCO_BOLETO
    return ""


def _base_bradesco(row: Mapping[str, Any]) -> Dict[str, str]:
    agora = datetime.now()
    data_geracao = _formata_data_cnab(row.get("data_geracao"), default=agora.strftime("%d%m%Y"))
    hora_geracao = _formata_hora_cnab(row.get("hora_geracao"), default=agora.strftime("%H%M%S"))

    agencia_combinada = _primeiro_valor_preenchido(
        row,
        "agencia_digito",
        "agencia_com_digito",
        "agencia_mais_digito",
    )
    agencia_numero = _primeiro_valor_preenchido(
        row,
        "agencia_empresa",
        "agencia",
    )
    agencia_origem = agencia_combinada if _texto(agencia_combinada) != "" else agencia_numero
    dv_agencia_origem = _primeiro_valor_preenchido(
        row,
        "dv_agencia_empresa",
        "digito_agencia",
        "dv_agencia",
    )
    agencia_empresa, dv_agencia_empresa = _cnab_num_dv(
        agencia_origem,
        dv_agencia_origem,
        largura_numero=5,
    )

    conta_combinada = _primeiro_valor_preenchido(
        row,
        "conta_digito",
        "conta_com_digito",
        "conta_mais_digito",
    )
    conta_numero = _primeiro_valor_preenchido(
        row,
        "conta_empresa",
        "conta",
    )
    conta_origem = conta_combinada if _texto(conta_combinada) != "" else conta_numero
    dv_conta_origem = _primeiro_valor_preenchido(
        row,
        "dv_conta_empresa",
        "digito_conta",
        "dv_conta",
    )
    conta_empresa, dv_conta_empresa = _cnab_num_dv(
        conta_origem,
        dv_conta_origem,
        largura_numero=12,
    )

    header_pix_raw = _texto(row.get("header_pix")).upper()
    header_pix = "PIX" if header_pix_raw == "PIX" else ""

    return {
        "banco": _cnab_num(row.get("banco"), 3, default="237"),
        "tipo_inscricao_empresa": _cnab_num(row.get("tipo_inscricao_empresa"), 1, default="2"),
        "numero_inscricao_empresa": _cnab_num(row.get("numero_inscricao_empresa"), 14),
        "convenio": _cnab_alfa(row.get("convenio"), 20),
        "agencia_empresa": agencia_empresa,
        "dv_agencia_empresa": dv_agencia_empresa,
        "conta_empresa": conta_empresa,
        "dv_conta_empresa": dv_conta_empresa,
        "dv_agencia_conta_empresa": _cnab_alfa(row.get("dv_agencia_conta_empresa"), 1),
        "nome_empresa": _cnab_alfa(row.get("nome_empresa"), 30),
        "nome_banco": _cnab_alfa(row.get("nome_banco"), 30, default="BANCO BRADESCO S.A"),
        "codigo_remessa_retorno": _cnab_num(row.get("codigo_remessa_retorno"), 1, default="1"),
        "data_geracao": data_geracao,
        "hora_geracao": hora_geracao,
        "nsa": _cnab_num(row.get("nsa"), 6, default="1"),
        "densidade": _cnab_num(row.get("densidade"), 5, default="0"),
        "header_pix": _cnab_alfa(header_pix, 3),
        "reservado_banco_arquivo": _cnab_alfa(row.get("reservado_banco_arquivo"), 17),
        "reservado_empresa_arquivo": _cnab_alfa(row.get("reservado_empresa_arquivo"), 20),
    }


def _primeiro_valor_preenchido(row: Mapping[str, Any], *chaves: str) -> Any:
    for chave in chaves:
        if chave not in row:
            continue
        valor = row.get(chave)
        if _texto(valor) != "":
            return valor
    return ""


def _cnab_num_dv(numero_valor: Any, dv_valor: Any, *, largura_numero: int) -> tuple[str, str]:
    texto_numero = _texto(numero_valor)
    texto_dv = _texto(dv_valor).upper()
    dv = texto_dv[:1] if texto_dv else ""

    numero_base = texto_numero
    match_numero_dv = re.match(r"^\s*([0-9]+)\s*[^0-9A-Za-z]+\s*([0-9A-Za-z])\s*$", texto_numero)
    if match_numero_dv:
        numero_base = match_numero_dv.group(1)
        if not dv:
            dv = match_numero_dv.group(2).upper()
    elif not dv:
        # Fallback para casos com separadores incomuns (ex.: hifen unicode) ou ruido textual.
        partes = re.findall(r"[0-9A-Za-z]+", texto_numero.upper())
        if len(partes) >= 2 and len(partes[-1]) == 1:
            candidato_numero = "".join(partes[:-1])
            if candidato_numero.isdigit():
                numero_base = candidato_numero
                dv = partes[-1]

    numero_digitos = _somente_digitos(numero_base)
    if not dv and len(numero_digitos) == largura_numero + 1:
        # Fallback para casos em que numero+dv chegam juntos sem separador (ex.: 05649).
        dv = numero_digitos[-1]
        numero_digitos = numero_digitos[:-1]

    return _cnab_num(numero_digitos, largura_numero), _cnab_alfa(dv, 1)


def _linha_header_arquivo_bradesco(base: Mapping[str, str]) -> str:
    campos = [
        base["banco"],
        "0000",
        "0",
        " " * 9,
        base["tipo_inscricao_empresa"],
        base["numero_inscricao_empresa"],
        base["convenio"],
        base["agencia_empresa"],
        base["dv_agencia_empresa"],
        base["conta_empresa"],
        base["dv_conta_empresa"],
        base["dv_agencia_conta_empresa"],
        base["nome_empresa"],
        base["nome_banco"],
        " " * 10,
        base["codigo_remessa_retorno"],
        base["data_geracao"],
        base["hora_geracao"],
        base["nsa"],
        "089",
        base["densidade"],
        base["header_pix"],
        base["reservado_banco_arquivo"],
        base["reservado_empresa_arquivo"],
        " " * 29,
    ]
    return _join_240(campos, "Header de arquivo Bradesco")


def _validar_header_pix_transferencia(base: Mapping[str, str]) -> None:
    header_pix = _texto(base.get("header_pix")).upper()
    if header_pix != "PIX":
        raise ConversionError(
            "BRADESCO_TRANSFERENCIA exige 'PIX' no header de arquivo (campo 172-174)."
        )


def _validar_posicao_pix_header_arquivo(linha: str) -> None:
    linha = _normalizar_linha_240(linha)
    if linha[171:174] != "PIX":
        raise ConversionError(
            "Header de arquivo invalido para BRADESCO_TRANSFERENCIA: campo 172-174 deve ser 'PIX'."
        )


def _linha_header_lote_transferencia(base: Mapping[str, str], lote: str, row: Mapping[str, Any]) -> str:
    campos = [
        base["banco"],
        lote,
        "1",
        "C",
        _cnab_num(row.get("tipo_servico"), 2, default="20"),
        # Logica mae: transferencia Bradesco sempre PIX.
        "45",
        "045",
        " ",
        base["tipo_inscricao_empresa"],
        base["numero_inscricao_empresa"],
        base["convenio"],
        base["agencia_empresa"],
        base["dv_agencia_empresa"],
        base["conta_empresa"],
        base["dv_conta_empresa"],
        base["dv_agencia_conta_empresa"],
        base["nome_empresa"],
        _cnab_alfa(row.get("mensagem_lote"), 40),
        _cnab_alfa(row.get("logradouro"), 30),
        _cnab_num(row.get("numero_local"), 5),
        _cnab_alfa(row.get("complemento_endereco"), 15),
        _cnab_alfa(row.get("cidade"), 20),
        _cnab_num(row.get("cep"), 5),
        _cnab_alfa(row.get("complemento_cep"), 3),
        _cnab_alfa(row.get("estado"), 2),
        _cnab_num(row.get("indicativo_forma_pagamento"), 2, default="1"),
        " " * 6,
        _cnab_alfa(row.get("ocorrencias_header_lote"), 10),
    ]
    return _join_240(campos, "Header de lote Transferencia")


def _linha_header_lote_boleto(base: Mapping[str, str], lote: str, row: Mapping[str, Any]) -> str:
    campos = [
        base["banco"],
        lote,
        "1",
        "C",
        _cnab_num(row.get("tipo_servico"), 2, default="30"),
        _cnab_num(row.get("forma_lancamento"), 2, default="30"),
        "040",
        " ",
        base["tipo_inscricao_empresa"],
        base["numero_inscricao_empresa"],
        base["convenio"],
        base["agencia_empresa"],
        base["dv_agencia_empresa"],
        base["conta_empresa"],
        base["dv_conta_empresa"],
        base["dv_agencia_conta_empresa"],
        base["nome_empresa"],
        _cnab_alfa(row.get("mensagem_lote"), 40),
        _cnab_alfa(row.get("logradouro"), 30),
        _cnab_num(row.get("numero_local"), 5),
        _cnab_alfa(row.get("complemento_endereco"), 15),
        _cnab_alfa(row.get("cidade"), 20),
        _cnab_num(row.get("cep"), 5),
        _cnab_alfa(row.get("complemento_cep"), 3),
        _cnab_alfa(row.get("estado"), 2),
        " " * 8,
        _cnab_alfa(row.get("ocorrencias_header_lote"), 10),
    ]
    return _join_240(campos, "Header de lote Boleto")


def _linha_segmento_a(base: Mapping[str, str], lote: str, row: Mapping[str, Any], sequencial: int) -> str:
    forma_iniciacao = _forma_iniciacao_pix(row)
    usa_dados_bancarios = forma_iniciacao == "05"
    banco_favorecido, agencia_favorecido, dv_agencia_favorecido, conta_favorecido, dv_conta_favorecido, dv_agencia_conta_favorecido = _dados_bancarios_favorecido_segmento_a(
        row, usa_dados_bancarios
    )
    codigo_finalidade_ted = _codigo_finalidade_ted(row)
    codigo_finalidade_complementar = _codigo_finalidade_complementar(row)

    campos = [
        base["banco"],
        lote,
        "3",
        _cnab_num(sequencial, 5),
        "A",
        _cnab_num(row.get("tipo_movimento"), 1, default="0"),
        _cnab_num(row.get("codigo_instrucao_movimento"), 2),
        "009",
        banco_favorecido,
        agencia_favorecido,
        dv_agencia_favorecido,
        conta_favorecido,
        dv_conta_favorecido,
        dv_agencia_conta_favorecido,
        _cnab_alfa(row.get("nome_favorecido"), 30),
        _numero_pagamento_segmento_a(row),
        _formata_data_cnab(row.get("data_pagamento")),
        _cnab_alfa(row.get("tipo_moeda"), 3, default="BRL"),
        _cnab_num(_numero_escalado(row.get("quantidade_moeda"), 5), 15),
        _cnab_num(_numero_escalado(row.get("valor_pagamento"), 2), 15),
        _cnab_alfa(row.get("nosso_numero"), 20),
        _formata_data_cnab(row.get("data_real_efetivacao"), default="00000000"),
        _cnab_num(_numero_escalado(row.get("valor_real_efetivacao"), 2), 15),
        _cnab_alfa(row.get("informacao2"), 40),
        _cnab_alfa(row.get("codigo_finalidade_doc"), 2),
        codigo_finalidade_ted,
        codigo_finalidade_complementar,
        " " * 3,
        _cnab_num(row.get("aviso_favorecido"), 1, default="0"),
        _cnab_alfa(row.get("ocorrencias"), 10),
    ]
    return _join_240(campos, "Segmento A")


def _linha_segmento_b(base: Mapping[str, str], lote: str, row: Mapping[str, Any], sequencial: int) -> str:
    forma_iniciacao = _forma_iniciacao_pix(row)
    tipo_inscricao_favorecido = _cnab_num(row.get("tipo_inscricao_favorecido"), 1, default="0")
    numero_inscricao_favorecido = _cnab_num(row.get("numero_inscricao_favorecido"), 14)
    _validar_identificacao_favorecido(
        row, forma_iniciacao, tipo_inscricao_favorecido, numero_inscricao_favorecido
    )
    informacao12 = _informacao12_pix(row, forma_iniciacao)

    campos = [
        base["banco"],
        lote,
        "3",
        _cnab_num(sequencial, 5),
        "B",
        f"{forma_iniciacao:<3}",
        tipo_inscricao_favorecido,
        numero_inscricao_favorecido,
        _cnab_alfa(row.get("informacao10"), 35),
        _cnab_alfa(row.get("informacao11"), 60),
        informacao12,
        _cnab_alfa(row.get("codigo_ug_centralizadora"), 6),
        _cnab_alfa(row.get("codigo_ispb"), 8),
    ]
    return _join_240(campos, "Segmento B")


def _forma_iniciacao_pix(row: Mapping[str, Any]) -> str:
    valor = _texto(row.get("forma_iniciacao"))
    digitos = _somente_digitos(valor)
    if digitos == "":
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: forma_iniciacao obrigatoria para PIX. Use 01, 02, 03, 04 ou 05."
        )
    forma = digitos[-2:].zfill(2)
    if forma not in {"01", "02", "03", "04", "05"}:
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: forma_iniciacao invalida ({valor!r}). Use 01, 02, 03, 04 ou 05."
        )
    return forma


def _dados_bancarios_favorecido_segmento_a(
    row: Mapping[str, Any], usa_dados_bancarios: bool
) -> tuple[str, str, str, str, str, str]:
    if not usa_dados_bancarios:
        # Quando a forma nao e 05, banco/agencia/conta do favorecido nao devem ser informados.
        return (" " * 3, " " * 5, " ", " " * 12, " ", " ")

    linha = _linha_planilha(row)
    banco_favorecido_raw = _texto(row.get("banco_favorecido"))
    agencia_favorecido_raw = _texto(row.get("agencia_favorecido"))
    dv_agencia_favorecido_raw = _texto(row.get("dv_agencia_favorecido")).upper()
    conta_favorecido_raw = _texto(row.get("conta_favorecido"))

    banco_favorecido = _cnab_num(banco_favorecido_raw, 3)
    agencia_favorecido_digitos = _somente_digitos(agencia_favorecido_raw)
    conta_favorecido = _cnab_num(conta_favorecido_raw, 12)

    if banco_favorecido == "000":
        raise ConversionError(
            f"Linha {linha}: banco_favorecido (coluna AL) obrigatorio na forma 05."
        )

    if agencia_favorecido_digitos == "":
        raise ConversionError(
            f"Linha {linha}: agencia_favorecido (coluna AM) obrigatoria na forma 05. Informe 4 ou 5 digitos."
        )
    if len(agencia_favorecido_digitos) not in {4, 5}:
        raise ConversionError(
            f"Linha {linha}: agencia_favorecido (coluna AM) invalida. Use 4 ou 5 digitos."
        )
    agencia_favorecido = agencia_favorecido_digitos.zfill(5)

    dv_agencia_favorecido_digitos = _somente_digitos(dv_agencia_favorecido_raw)
    if dv_agencia_favorecido_digitos == "":
        raise ConversionError(
            f"Linha {linha}: dv_agencia_favorecido (coluna AN) obrigatorio na forma 05. Informe 1 digito."
        )
    if len(dv_agencia_favorecido_digitos) != 1:
        raise ConversionError(
            f"Linha {linha}: dv_agencia_favorecido (coluna AN) invalido. Use somente 1 digito."
        )
    dv_agencia_favorecido = dv_agencia_favorecido_digitos

    if conta_favorecido == "000000000000":
        raise ConversionError(
            f"Linha {linha}: conta_favorecido (coluna AO) obrigatoria na forma 05."
        )

    return (
        banco_favorecido,
        agencia_favorecido,
        dv_agencia_favorecido,
        conta_favorecido,
        _cnab_alfa(row.get("dv_conta_favorecido"), 1),
        _cnab_alfa(row.get("dv_agencia_conta_favorecido"), 1),
    )


def _validar_identificacao_favorecido(
    row: Mapping[str, Any], forma_iniciacao: str, tipo_inscricao_favorecido: str, numero_inscricao_favorecido: str
) -> None:
    if tipo_inscricao_favorecido not in {"1", "2"}:
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: tipo de inscricao do favorecido invalido. Use 1 para CPF ou 2 para CNPJ."
        )
    if numero_inscricao_favorecido == "00000000000000":
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: numero de inscricao do favorecido obrigatorio para PIX (segmento B, posicoes 19-32)."
        )


def _informacao12_pix(row: Mapping[str, Any], forma_iniciacao: str) -> str:
    chave_pix = _texto(row.get("informacao12"))
    if forma_iniciacao in {"01", "02", "04"}:
        if chave_pix == "":
            raise ConversionError(
                f"Linha {_linha_planilha(row)}: forma_iniciacao {forma_iniciacao} exige chave PIX em informacao12 (posicoes 128-226 do segmento B)."
            )
        return _cnab_alfa(chave_pix, 99)

    if forma_iniciacao == "03":
        # Para forma 03 a chave nao e obrigatoria, mas se vier preenchida pode ser enviada.
        if chave_pix != "":
            conteudo = ("   " + chave_pix).replace("\r", " ").replace("\n", " ")
            return f"{conteudo[:99]:<99}"
        return " " * 99

    tipo_conta_raw = _texto(row.get("tipo_conta_recebedor"))
    if tipo_conta_raw == "":
        complementar = _texto(row.get("codigo_finalidade_complementar")).upper()
        if complementar == "CC":
            tipo_conta_raw = "01"
        elif complementar == "PP":
            tipo_conta_raw = "03"

    tipo_conta = _somente_digitos(tipo_conta_raw)
    if tipo_conta == "":
        tipo_conta = _somente_digitos(chave_pix)
    tipo_conta = tipo_conta[-2:].zfill(2) if tipo_conta != "" else ""

    if tipo_conta not in {"01", "02", "03"}:
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: forma de iniciacao 05 exige tipo de conta do recebedor em informacao12 (01=conta corrente, 02=conta pagamento, 03=conta poupanca)."
        )
    return f"{tipo_conta}{' ' * 97}"


def _validar_modalidade_pix_transferencia(row: Mapping[str, Any]) -> None:
    numero_pagamento = _texto(_numero_pagamento(row))
    if numero_pagamento == "":
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: numero_pagamento obrigatorio (segmento A, posicoes 74-93)."
        )

    forma_lancamento_raw = _texto(row.get("forma_lancamento"))
    forma_lancamento = _somente_digitos(forma_lancamento_raw)
    if forma_lancamento != "" and forma_lancamento[-2:].zfill(2) != "45":
        raise ConversionError(
            f"Linha {_linha_planilha(row)}: forma_lancamento {forma_lancamento_raw!r} invalida para BRADESCO_TRANSFERENCIA. Informe 45 (PIX Transferencia)."
        )


def _numero_pagamento(row: Mapping[str, Any]) -> Any:
    return _primeiro_valor_preenchido(row, "numero_pagamento", "seu_numero")


def _numero_pagamento_segmento_a(row: Mapping[str, Any]) -> str:
    valor = _texto(_numero_pagamento(row))
    return f"{valor[:20]:>20}"


def _linha_planilha(row: Mapping[str, Any]) -> str:
    linha = row.get("__linha_planilha")
    if linha is None:
        return "?"
    return str(linha)


def _codigo_finalidade_ted(row: Mapping[str, Any]) -> str:
    valor = _texto(row.get("codigo_finalidade_ted"))
    if valor == "":
        return "00010"

    digitos = _somente_digitos(valor)
    if digitos == "":
        raise ConversionError(
            "Codigo finalidade TED invalido. Informe um codigo numerico de 5 digitos (ex.: 00010)."
        )
    return digitos[-5:].zfill(5)


def _codigo_finalidade_complementar(row: Mapping[str, Any]) -> str:
    valor = _texto(row.get("codigo_finalidade_complementar")).upper()
    if valor == "":
        return "CC"

    if valor not in {"CC", "PP"}:
        raise ConversionError(
            "Codigo finalidade complementar invalido. Use 'CC' (corrente) ou 'PP' (poupanca)."
        )
    return valor


def _linha_segmento_j(base: Mapping[str, str], lote: str, row: Mapping[str, Any], sequencial: int) -> str:
    campos = [
        base["banco"],
        lote,
        "3",
        _cnab_num(sequencial, 5),
        "J",
        _cnab_num(row.get("tipo_movimento"), 1, default="0"),
        _cnab_num(row.get("codigo_instrucao_movimento"), 2),
        _cnab_num(row.get("codigo_barras"), 44),
        _cnab_alfa(row.get("nome_beneficiario"), 30),
        _formata_data_cnab(row.get("data_vencimento")),
        _cnab_num(_numero_escalado(row.get("valor_titulo"), 2), 15),
        _cnab_num(_numero_escalado(row.get("desconto_abatimento"), 2), 15),
        _cnab_num(_numero_escalado(row.get("acrescimos"), 2), 15),
        _formata_data_cnab(row.get("data_pagamento")),
        _cnab_num(_numero_escalado(row.get("valor_pagamento"), 2), 15),
        _cnab_num(_numero_escalado(row.get("quantidade_moeda"), 5), 15),
        _cnab_alfa(row.get("referencia_pagador"), 20),
        _cnab_alfa(row.get("nosso_numero"), 20),
        _cnab_num(row.get("codigo_moeda"), 2, default="9"),
        " " * 6,
        _cnab_alfa(row.get("ocorrencias"), 10),
    ]
    return _join_240(campos, "Segmento J")


def _linha_segmento_j52(base: Mapping[str, str], lote: str, row: Mapping[str, Any], sequencial: int) -> str:
    chave_pix = _texto(row.get("chave_pagamento"))
    txid = _texto(row.get("txid"))
    usa_pix = chave_pix != "" or txid != ""

    campos = [
        base["banco"],
        lote,
        "3",
        _cnab_num(sequencial, 5),
        "J",
        " ",
        _cnab_num(row.get("codigo_movimento_remessa"), 2, default="00"),
        "52",
        _cnab_num(row.get("tipo_inscricao_pagador"), 1, default="0"),
        _cnab_num(row.get("numero_inscricao_pagador"), 15),
        _cnab_alfa(row.get("nome_pagador"), 40),
        _cnab_num(row.get("tipo_inscricao_beneficiario"), 1, default="0"),
        _cnab_num(row.get("numero_inscricao_beneficiario"), 15),
        _cnab_alfa(row.get("nome_beneficiario_j52"), 40),
    ]

    if usa_pix:
        campos.extend(
            [
                _cnab_alfa(chave_pix, 79),
                _cnab_alfa(txid, 30),
            ]
        )
    else:
        campos.extend(
            [
                _cnab_num(row.get("tipo_inscricao_sacador"), 1, default="0"),
                _cnab_num(row.get("numero_inscricao_sacador"), 15),
                _cnab_alfa(row.get("nome_sacador"), 40),
                " " * 53,
            ]
        )
    return _join_240(campos, "Segmento J-52")


def _linha_trailer_lote(
    *,
    base: Mapping[str, str],
    lote: str,
    qtd_registros_lote: int,
    soma_valores: int,
    soma_qtd_moeda: int,
    numero_aviso_debito: Any,
) -> str:
    campos = [
        base["banco"],
        lote,
        "5",
        " " * 9,
        _cnab_num(qtd_registros_lote, 6),
        _cnab_num(soma_valores, 18),
        _cnab_num(soma_qtd_moeda, 18),
        _cnab_num(numero_aviso_debito, 6),
        " " * 165,
        " " * 10,
    ]
    return _join_240(campos, "Trailer de lote")


def _linha_trailer_arquivo(
    *,
    base: Mapping[str, str],
    qtd_lotes: int,
    qtd_registros_arquivo: int,
    qtd_contas_conciliacao: Any,
) -> str:
    campos = [
        base["banco"],
        "9999",
        "9",
        " " * 9,
        _cnab_num(qtd_lotes, 6),
        _cnab_num(qtd_registros_arquivo, 6),
        _cnab_num(qtd_contas_conciliacao, 6),
        " " * 205,
    ]
    return _join_240(campos, "Trailer de arquivo")


def _join_240(campos: List[str], nome_registro: str) -> str:
    linha = "".join(campos)
    if len(linha) != 240:
        raise ConversionError(
            f"{nome_registro} com tamanho invalido ({len(linha)}). Esperado: 240."
        )
    return linha


def _cnab_alfa(valor: Any, largura: int, default: str = "") -> str:
    texto = _texto(valor)
    if texto == "":
        texto = default
    texto = texto.replace("\r", " ").replace("\n", " ")
    return f"{texto[:largura]:<{largura}}"


def _cnab_num(valor: Any, largura: int, default: str = "0") -> str:
    if isinstance(valor, int):
        digitos = str(valor)
    else:
        digitos = _somente_digitos(valor)
    if digitos == "":
        digitos = _somente_digitos(default) or "0"
    return digitos[-largura:].zfill(largura)


def _numero_escalado(valor: Any, casas_decimais: int) -> int:
    texto = _texto(valor)
    if texto == "":
        return 0
    texto = _normalizar_decimal(texto)
    try:
        decimal = Decimal(texto)
    except (InvalidOperation, ValueError) as exc:
        raise ConversionError(f"Valor monetario invalido: {valor!r}") from exc
    fator = Decimal(10) ** casas_decimais
    return int((decimal * fator).quantize(Decimal("1")))


def _formata_data_cnab(valor: Any, default: str = "00000000") -> str:
    texto_original = _texto(valor).strip()
    if texto_original == "":
        return _somente_digitos(default).zfill(8)[-8:]

    # Se vier como numero serial do Excel.
    if re.fullmatch(r"\d+(\.0+)?", texto_original):
        serial_excel = int(float(texto_original))
        if 20000 <= serial_excel <= 90000:
            data_excel = datetime(1899, 12, 30) + timedelta(days=serial_excel)
            return data_excel.strftime("%d%m%Y")

    digitos = _somente_digitos(texto_original)

    # yyyy-mm-dd, yyyy/mm/dd, yyyy.mm.dd
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            dt = datetime.strptime(texto_original, fmt)
            return dt.strftime("%d%m%Y")
        except ValueError:
            pass

    # dd-mm-yyyy, dd/mm/yyyy, dd.mm.yyyy
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            dt = datetime.strptime(texto_original, fmt)
            return dt.strftime("%d%m%Y")
        except ValueError:
            pass

    # yyyymmdd
    if len(digitos) == 8:
        try:
            dt = datetime.strptime(digitos, "%Y%m%d")
            return dt.strftime("%d%m%Y")
        except ValueError:
            pass

    # ddmmaaaa
    if len(digitos) == 8:
        try:
            dt = datetime.strptime(digitos, "%d%m%Y")
            return dt.strftime("%d%m%Y")
        except ValueError:
            pass

    raise ConversionError(f"Data invalida para CNAB: {valor!r}. Use dd/mm/aaaa.")


def _formata_hora_cnab(valor: Any, default: str = "000000") -> str:
    texto = _somente_digitos(_texto(valor))
    if texto == "":
        texto = _somente_digitos(default)
    return texto[-6:].zfill(6)


def _formatar_linha_retorno(row: Mapping[str, Any]) -> str:
    linha = (
        f"{_texto(row.get('ano_mes'))[:6]:<6}"
        f"{_somente_digitos(row.get('orgao')).zfill(3)}"
        f"{_texto(row.get('matricula'))[:8]:<8}"
        f"{_somente_digitos(row.get('consignataria')).zfill(6)}"
        f"{_valor_em_centavos(row.get('valor_parcela'), 11)}"
        "00"
        f"{_somente_digitos(row.get('cpf')).zfill(11)}"
        f"{_texto(row.get('contrato')).zfill(20)}"
        f"{_texto(row.get('nome_servidor'))[:27]:<27}"
    )
    if len(linha) != 94:
        raise ConversionError("Linha RETORNO com tamanho invalido. Esperado: 94.")
    return linha


def _formatar_linha_variacao(row: Mapping[str, Any]) -> str:
    tipo_ajuste = _texto(row.get("tipo_ajuste")).upper()[:1]
    categoria_ajuste = _texto(row.get("categoria_ajuste"))[:1]
    valor_parcela = (
        _valor_em_centavos(row.get("valor_parcela"), 7)
        if tipo_ajuste != "E"
        else "0000000"
    )

    linha = (
        f"{_numero_inteiro(row.get('consignataria'), 6)}"
        f"{_texto(row.get('nome_consignataria'))[:20]:<20}"
        f"{_texto(row.get('instituicao'))[:20]:<20}"
        f"{_numero_inteiro(row.get('contrato'), 15)}"
        f"{_texto(row.get('nome_servidor'))[:30]:<30}"
        f"{_somente_digitos(row.get('cpf')).zfill(11)}"
        f"{_numero_inteiro(row.get('orgao'), 3)}"
        f"{_numero_inteiro(row.get('matricula'), 8)}"
        f"{tipo_ajuste}"
        f"{categoria_ajuste}"
        f"{_texto(row.get('data_inicial')).zfill(8)}"
        f"{_valor_em_centavos(row.get('valor_total'), 10)}"
        f"{_numero_inteiro(row.get('qtd_parcelas'), 3)}"
        f"{valor_parcela}"
        f"{_numero_inteiro(row.get('proxima_parcela'), 3)}"
    )
    if len(linha) != 146:
        raise ConversionError("Linha VARIACAO com tamanho invalido. Esperado: 146.")
    return linha


def _somente_digitos(valor: Any) -> str:
    return re.sub(r"\D", "", _texto(valor))


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d%m%Y")
    if isinstance(valor, date):
        return valor.strftime("%d%m%Y")
    if isinstance(valor, float):
        if math.isnan(valor):
            return ""
        if valor.is_integer():
            return str(int(valor))
        return format(valor, "f").rstrip("0").rstrip(".")
    texto = str(valor).strip()
    if texto.endswith(".0") and texto.replace(".", "", 1).isdigit():
        return texto[:-2]
    return texto


def _numero_inteiro(valor: Any, largura: int) -> str:
    texto = _texto(valor)
    if texto == "":
        numero = 0
    else:
        texto = _normalizar_decimal(texto)
        try:
            numero = int(Decimal(texto))
        except (InvalidOperation, ValueError) as exc:
            raise ConversionError(f"Valor inteiro invalido: {valor!r}") from exc
    return f"{numero:0{largura}d}"


def _valor_em_centavos(valor: Any, largura: int) -> str:
    texto = _texto(valor)
    if texto == "":
        centavos = 0
    else:
        texto = _normalizar_decimal(texto)
        try:
            centavos = int(Decimal(texto) * 100)
        except (InvalidOperation, ValueError) as exc:
            raise ConversionError(f"Valor monetario invalido: {valor!r}") from exc
    return f"{centavos:0{largura}d}"


def _normalizar_decimal(texto: str) -> str:
    limpo = re.sub(r"[^\d,.\-]", "", texto)
    if "," in limpo and "." in limpo:
        limpo = limpo.replace(".", "").replace(",", ".")
    elif "," in limpo:
        limpo = limpo.replace(",", ".")
    return limpo
